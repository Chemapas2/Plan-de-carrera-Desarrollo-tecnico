import io
import json
import math
import statistics
import tempfile
from pathlib import Path

import pandas as pd
import streamlit as st
from openpyxl import load_workbook

st.set_page_config(
    page_title="Senior Technical Consultant Selector",
    page_icon="🏅",
    layout="wide",
)

# =========================================================
# CONFIGURACIÓN DEL MODELO
# =========================================================

CRITICAL_INDICATORS = {
    "Nutrición en general": 1.00,
    "Conocimiento de productos": 1.00,
    "Patología metabólica": 1.10,
    "Aditivos alternativos": 1.00,
    "Bioseguridad": 1.00,
    "Arranques": 1.10,
    "Animales en producción": 1.10,
    "Tratamiento de datos": 1.00,
    "Informes": 0.95,
    "Ingles": 0.90,
    "Manejo herramientas/programas": 0.95,
}

TRANSFER_INDICATORS = [
    "Conocimiento de productos",
    "Tratamiento de datos",
    "Tratamiento de textos",
    "Informes",
    "Ingles",
    "Manejo herramientas/programas",
]

RANKING_WEIGHTS = {
    "global_performance": 0.35,
    "balance": 0.20,
    "critical": 0.20,
    "transfer": 0.15,
    "team_advantage": 0.10,
}

CRITERIA_TABLE = pd.DataFrame(
    [
        {
            "Criterio": "Rendimiento técnico global",
            "Peso": "35%",
            "Qué mide": "Nivel técnico total del candidato dentro de la propia herramienta.",
            "Cómo se calcula": "Combinación de la comparación global vs objetivo y de la comparación global vs máximo de referencia.",
        },
        {
            "Criterio": "Equilibrio entre los 4 troncos",
            "Peso": "20%",
            "Qué mide": "Si el perfil es completo y consistente o está demasiado descompensado.",
            "Cómo se calcula": "Media de los 4 troncos vs objetivo + refuerzo del tronco más débil para penalizar desequilibrios.",
        },
        {
            "Criterio": "Indicadores críticos para senior",
            "Peso": "20%",
            "Qué mide": "Fortaleza en las materias más relevantes para actuar como referente senior.",
            "Cómo se calcula": "Promedio ponderado de indicadores críticos vs objetivo y vs máximo, con penalización por carencias graves.",
        },
        {
            "Criterio": "Capacidad de transferencia y formación",
            "Peso": "15%",
            "Qué mide": "Potencial para formar, estructurar criterio y ayudar a crecer al resto del equipo.",
            "Cómo se calcula": "Subconjunto de indicadores ligados a comunicación, informes, datos, herramientas e inglés.",
        },
        {
            "Criterio": "Ventaja respecto a la media del equipo",
            "Peso": "10%",
            "Qué mide": "Cuánto destaca realmente el candidato frente al estándar medio del equipo.",
            "Cómo se calcula": "Comparación global vs BBDD + porcentaje de indicadores por encima de la media BBDD.",
        },
    ]
)

CRITERIA_TEXT = """
**Regla clave de justicia:** la app no elige automáticamente a quien tiene la mayor nota global.
Elige al perfil más adecuado para actuar como **consultor senior**, es decir, un técnico sólido,
equilibrado, fuerte en los indicadores críticos y con capacidad de referencia y de formación.

**Antes del ranking final hay un filtro mínimo de elegibilidad senior.**
Un candidato solo se considera “Apto ahora” si cumple los cuatro requisitos:
1. Está **por encima del objetivo global**.
2. No tiene ningún tronco **claramente débil**.
3. No acumula carencias graves en los **indicadores críticos**.
4. Tiene una base suficiente en **transferencia/formación**.

Si nadie cumple ese estándar, la app puede concluir que **todavía no hay un senior claro**.
"""


POSSIBLE_QUESTIONS = [
    "¿El evaluado podría ser formado en áreas muy específicas o exigentes?",
    "Además de las áreas propuestas, ¿qué otras tienen más posibilidades de éxito?",
    "¿Está preparado para asumir mayor responsabilidad técnica a corto plazo?",
    "¿Qué debilidades limitan más su evolución en los próximos 12 meses?",
    "¿Qué fortalezas son más diferenciales respecto al resto del equipo?",
    "¿En qué tronco debería centrarse primero el plan de mejora?",
    "¿Qué indicadores críticos requieren una intervención formativa inmediata?",
    "¿Tiene potencial real para convertirse en consultor senior en el futuro?",
    "¿Qué tipo de formación le encaja mejor: base, especializada o avanzada?",
    "¿Sería recomendable exponerlo a proyectos transversales o de mayor complejidad?",
    "¿Puede actuar ya como referente parcial en alguna materia concreta?",
    "¿Qué gap le separa con más claridad del perfil senior deseado?",
    "¿Qué tipo de acompañamiento interno necesitaría para acelerar su desarrollo?",
    "¿Qué fortalezas conviene consolidar para convertirlas en especialización?",
    "¿Hay riesgo de sobreestimar su perfil por desequilibrios entre troncos?",
    "¿Qué impacto tendría priorizar su formación en herramientas y análisis de datos?",
    "¿Qué combinación de formación interna y externa parece más eficiente para este técnico?",
    "¿Está más cerca de un perfil especialista o de un perfil generalista sólido?",
    "¿Qué argumentos objetivos justifican que no sea el elegido ahora mismo?",
    "¿Cuál sería el siguiente paso más razonable en su itinerario profesional?",
]

# =========================================================
# UTILIDADES
# =========================================================


def initialize_app_state():
    if "uploader_key" not in st.session_state:
        st.session_state["uploader_key"] = 0
    if "qa_items" not in st.session_state:
        st.session_state["qa_items"] = []
    if "report_text" not in st.session_state:
        st.session_state["report_text"] = ""
    if "question_selector" not in st.session_state:
        st.session_state["question_selector"] = POSSIBLE_QUESTIONS[0]


def reset_app_state():
    st.session_state["uploader_key"] += 1
    st.session_state["qa_items"] = []
    st.session_state["report_text"] = ""
    st.session_state["question_selector"] = POSSIBLE_QUESTIONS[0]

def safe_float(value):
    if value in (None, ""):
        return None
    try:
        return float(value)
    except Exception:
        return None


def ratio_to_score(value, stretch=1.20):
    value = safe_float(value)
    if value is None or math.isnan(value):
        return 0.0
    value = max(0.0, min(value, stretch))
    return (value / stretch) * 100.0


def pct(value):
    value = safe_float(value)
    if value is None:
        return None
    return round(value * 100.0, 1)


def format_pct(value):
    value = pct(value)
    return "-" if value is None else f"{value:.1f}%"


def strongest_and_weakest_trunk(trunks_df: pd.DataFrame):
    if trunks_df.empty:
        return "-", "-"
    tmp = trunks_df.copy().dropna(subset=["vs_goal"])
    if tmp.empty:
        return "-", "-"
    strongest = tmp.sort_values("vs_goal", ascending=False).iloc[0]["tronco"]
    weakest = tmp.sort_values("vs_goal", ascending=True).iloc[0]["tronco"]
    return strongest, weakest


def build_indicator_frame(ref_ws, eval_ws) -> pd.DataFrame:
    rows = []
    current_area = None

    for ref_row, eval_row in zip(range(4, 29), range(9, 34)):
        area = ref_ws[f"B{ref_row}"].value
        if area:
            current_area = str(area).strip()

        indicator = str(ref_ws[f"C{ref_row}"].value).strip()
        weight = safe_float(ref_ws[f"D{ref_row}"].value) or 0.0
        objective_raw = safe_float(ref_ws[f"E{ref_row}"].value) or 0.0
        objective_weighted = safe_float(ref_ws[f"F{ref_row}"].value)
        max_weighted = safe_float(ref_ws[f"G{ref_row}"].value)
        bbdd_raw = safe_float(ref_ws[f"H{ref_row}"].value)
        bbdd_weighted = safe_float(ref_ws[f"I{ref_row}"].value)

        raw_score = safe_float(eval_ws[f"D{eval_row}"].value)
        weighted_score = None if raw_score is None else raw_score * weight

        if objective_weighted is None:
            objective_weighted = weight * objective_raw
        if max_weighted is None:
            max_weighted = weight * 4
        if bbdd_weighted is None and bbdd_raw is not None:
            bbdd_weighted = weight * bbdd_raw

        vs_goal = None if not objective_weighted else weighted_score / objective_weighted if weighted_score is not None else None
        vs_max = None if not max_weighted else weighted_score / max_weighted if weighted_score is not None else None
        vs_bbdd = None if not bbdd_weighted else weighted_score / bbdd_weighted if weighted_score is not None else None

        rows.append(
            {
                "tronco": current_area,
                "indicator": indicator,
                "weight": weight,
                "score_raw": raw_score,
                "score_weighted": weighted_score,
                "objective_raw": objective_raw,
                "objective_weighted": objective_weighted,
                "max_weighted": max_weighted,
                "bbdd_raw": bbdd_raw,
                "bbdd_weighted": bbdd_weighted,
                "vs_goal": vs_goal,
                "vs_max": vs_max,
                "vs_bbdd": vs_bbdd,
            }
        )

    return pd.DataFrame(rows)


def summarise_trunks(indicators_df: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for trunk, grp in indicators_df.groupby("tronco", dropna=False):
        rows.append(
            {
                "tronco": trunk,
                "score_raw_total": grp["score_raw"].sum(skipna=True),
                "score_weighted_total": grp["score_weighted"].sum(skipna=True),
                "objective_weighted_total": grp["objective_weighted"].sum(skipna=True),
                "max_weighted_total": grp["max_weighted"].sum(skipna=True),
                "bbdd_weighted_total": grp["bbdd_weighted"].sum(skipna=True),
                "avg_raw": grp["score_raw"].mean(skipna=True),
                "vs_goal": (grp["score_weighted"].sum(skipna=True) / grp["objective_weighted"].sum(skipna=True))
                if grp["objective_weighted"].sum(skipna=True) else None,
                "vs_max": (grp["score_weighted"].sum(skipna=True) / grp["max_weighted"].sum(skipna=True))
                if grp["max_weighted"].sum(skipna=True) else None,
                "vs_bbdd": (grp["score_weighted"].sum(skipna=True) / grp["bbdd_weighted"].sum(skipna=True))
                if grp["bbdd_weighted"].sum(skipna=True) else None,
            }
        )
    return pd.DataFrame(rows)


def summarise_global(indicators_df: pd.DataFrame) -> dict:
    score_raw_total = indicators_df["score_raw"].sum(skipna=True)
    score_weighted_total = indicators_df["score_weighted"].sum(skipna=True)
    objective_weighted_total = indicators_df["objective_weighted"].sum(skipna=True)
    max_weighted_total = indicators_df["max_weighted"].sum(skipna=True)
    bbdd_weighted_total = indicators_df["bbdd_weighted"].sum(skipna=True)
    avg_raw = indicators_df["score_raw"].mean(skipna=True)
    bbdd_raw_avg = indicators_df["bbdd_raw"].mean(skipna=True)

    return {
        "score_raw_total": score_raw_total,
        "score_weighted_total": score_weighted_total,
        "objective_weighted_total": objective_weighted_total,
        "max_weighted_total": max_weighted_total,
        "bbdd_weighted_total": bbdd_weighted_total,
        "avg_raw": avg_raw,
        "bbdd_raw_avg": bbdd_raw_avg,
        "vs_goal": (score_weighted_total / objective_weighted_total) if objective_weighted_total else None,
        "vs_max": (score_weighted_total / max_weighted_total) if max_weighted_total else None,
        "vs_bbdd": (score_weighted_total / bbdd_weighted_total) if bbdd_weighted_total else None,
    }


def infer_level(avg_raw: float) -> str:
    if avg_raw is None or math.isnan(avg_raw):
        return "-"
    if avg_raw < 1:
        return "BÁSICO"
    if avg_raw < 2:
        return "CONTROLA"
    if avg_raw < 3:
        return "SUPERA"
    if avg_raw < 4:
        return "CERTIFICADO"
    if avg_raw < 5:
        return "EXCELENTE"
    if avg_raw < 6:
        return "MASTER"
    return "MÁXIMO"


def parse_candidate(uploaded_file):
    suffix = Path(uploaded_file.name).suffix or ".xlsm"

    with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
        tmp.write(uploaded_file.getvalue())
        temp_path = tmp.name

    wb = load_workbook(temp_path, data_only=True, keep_vba=True)
    required = {"REFERENCIAS", "EVALUACION"}
    if not required.issubset(set(wb.sheetnames)):
        raise ValueError(
            f"El archivo {uploaded_file.name} no contiene las hojas mínimas esperadas: {', '.join(sorted(required))}."
        )

    ref_ws = wb["REFERENCIAS"]
    eval_ws = wb["EVALUACION"]
    mod_ws = wb["MODULO"] if "MODULO" in wb.sheetnames else None

    name = eval_ws["C2"].value or (mod_ws["H14"].value if mod_ws else None) or uploaded_file.name
    species = eval_ws["C4"].value or (mod_ws["H15"].value if mod_ws else None) or "-"
    date_value = eval_ws["C6"].value or (mod_ws["H16"].value if mod_ws else None)

    indicators_df = build_indicator_frame(ref_ws, eval_ws)
    if indicators_df["score_raw"].notna().sum() < 20:
        raise ValueError(
            f"El archivo {uploaded_file.name} no parece tener suficientes puntuaciones cargadas. "
            "Ábrelo en Excel, recalcula/guarda y vuelve a subirlo."
        )

    trunks_df = summarise_trunks(indicators_df)
    global_summary = summarise_global(indicators_df)
    strongest_trunk, weakest_trunk = strongest_and_weakest_trunk(trunks_df)

    return {
        "filename": uploaded_file.name,
        "name": str(name).strip(),
        "species": str(species).strip() if species is not None else "-",
        "date": date_value,
        "indicators": indicators_df,
        "trunks": trunks_df,
        "global": global_summary,
        "global_level": infer_level(global_summary["avg_raw"]),
        "strongest_trunk": strongest_trunk,
        "weakest_trunk": weakest_trunk,
    }


def score_candidate(candidate: dict):
    indicators_df = candidate["indicators"].copy()
    trunks_df = candidate["trunks"].copy()
    global_summary = candidate["global"]

    global_score = (
        0.60 * ratio_to_score(global_summary["vs_goal"], stretch=1.15)
        + 0.40 * ratio_to_score(global_summary["vs_max"], stretch=1.00)
    )

    trunk_goal_values = [v for v in trunks_df["vs_goal"].tolist() if v is not None and not pd.isna(v)]
    avg_trunk_goal = statistics.mean(trunk_goal_values) if trunk_goal_values else 0.0
    min_trunk_goal = min(trunk_goal_values) if trunk_goal_values else 0.0
    balance_score = (
        0.60 * ratio_to_score(avg_trunk_goal, stretch=1.10)
        + 0.40 * ratio_to_score(min_trunk_goal, stretch=1.00)
    )

    critical_df = indicators_df[indicators_df["indicator"].isin(CRITICAL_INDICATORS.keys())].copy()
    if critical_df.empty:
        critical_score = 0.0
        weak_critical_count = 999
    else:
        critical_df["priority_weight"] = critical_df["indicator"].map(CRITICAL_INDICATORS).fillna(1.0)
        critical_df["goal_score"] = critical_df["vs_goal"].apply(lambda x: ratio_to_score(x, stretch=1.15))
        critical_df["max_score"] = critical_df["vs_max"].apply(lambda x: ratio_to_score(x, stretch=1.00))
        weighted_goal = (critical_df["goal_score"] * critical_df["priority_weight"]).sum() / critical_df["priority_weight"].sum()
        weighted_max = (critical_df["max_score"] * critical_df["priority_weight"]).sum() / critical_df["priority_weight"].sum()
        weak_critical_count = int((critical_df["vs_goal"] < 0.85).fillna(False).sum())
        penalty = min(24, weak_critical_count * 8)
        critical_score = max(0.0, (0.70 * weighted_goal + 0.30 * weighted_max) - penalty)

    transfer_df = indicators_df[indicators_df["indicator"].isin(TRANSFER_INDICATORS)].copy()
    if transfer_df.empty:
        transfer_score = 0.0
    else:
        transfer_goal = transfer_df["vs_goal"].apply(lambda x: ratio_to_score(x, stretch=1.10)).mean()
        transfer_bbdd = transfer_df["vs_bbdd"].apply(lambda x: ratio_to_score(x, stretch=1.20)).mean()
        transfer_score = 0.70 * transfer_goal + 0.30 * transfer_bbdd

    above_bbdd_pct = float((indicators_df["vs_bbdd"] >= 1.0).fillna(False).mean() * 100.0)
    team_advantage_score = (
        0.70 * ratio_to_score(global_summary["vs_bbdd"], stretch=1.25)
        + 0.30 * above_bbdd_pct
    )

    senior_score = (
        RANKING_WEIGHTS["global_performance"] * global_score
        + RANKING_WEIGHTS["balance"] * balance_score
        + RANKING_WEIGHTS["critical"] * critical_score
        + RANKING_WEIGHTS["transfer"] * transfer_score
        + RANKING_WEIGHTS["team_advantage"] * team_advantage_score
    )

    min_trunk = min_trunk_goal if trunk_goal_values else 0.0
    eligibility_checks = {
        "global_vs_goal": (global_summary["vs_goal"] is not None and global_summary["vs_goal"] >= 1.00),
        "weakest_trunk": (min_trunk >= 0.90),
        "critical_gaps": (weak_critical_count <= 1),
        "transfer_floor": (transfer_score >= 55.0),
    }
    eligible_now = all(eligibility_checks.values())

    return {
        "component_global": round(global_score, 1),
        "component_balance": round(balance_score, 1),
        "component_critical": round(critical_score, 1),
        "component_transfer": round(transfer_score, 1),
        "component_team_advantage": round(team_advantage_score, 1),
        "senior_score": round(senior_score, 1),
        "eligible_now": eligible_now,
        "eligibility_checks": eligibility_checks,
        "weak_critical_count": weak_critical_count,
        "above_bbdd_pct": round(above_bbdd_pct, 1),
    }


def top_strengths(indicators_df: pd.DataFrame, top_n=3):
    tmp = indicators_df.copy()
    tmp["composite"] = (
        tmp["vs_goal"].fillna(0) * 0.60
        + tmp["vs_max"].fillna(0) * 0.25
        + tmp["vs_bbdd"].fillna(0) * 0.15
    )
    top = tmp.sort_values("composite", ascending=False).head(top_n)
    return top[["indicator", "tronco", "score_raw", "vs_goal", "vs_bbdd"]].to_dict("records")


def main_gaps(indicators_df: pd.DataFrame, top_n=3):
    tmp = indicators_df.copy()
    tmp["is_critical"] = tmp["indicator"].isin(CRITICAL_INDICATORS.keys())
    tmp["priority"] = tmp["is_critical"].astype(int)
    tmp["composite"] = (
        tmp["vs_goal"].fillna(0) * 0.70
        + tmp["vs_bbdd"].fillna(0) * 0.30
    )
    tmp = tmp.sort_values(["priority", "composite"], ascending=[False, True])
    gaps = tmp.head(top_n)
    return gaps[["indicator", "tronco", "score_raw", "vs_goal", "vs_bbdd"]].to_dict("records")


def candidate_argument(candidate: dict, score_data: dict, selected_name: str = None):
    global_summary = candidate["global"]
    strengths = top_strengths(candidate["indicators"], top_n=3)
    gaps = main_gaps(candidate["indicators"], top_n=3)

    strengths_txt = "; ".join(
        [f"{x['indicator']} ({format_pct(x['vs_goal'])} vs objetivo)" for x in strengths]
    ) or "sin fortalezas destacadas"
    gaps_txt = "; ".join(
        [f"{x['indicator']} ({format_pct(x['vs_goal'])} vs objetivo)" for x in gaps]
    ) or "sin gaps relevantes"

    if candidate["name"] == selected_name and score_data["eligible_now"]:
        return (
            f"**Elección propuesta para consultor senior.** "
            f"Obtiene el mejor Senior Score ({score_data['senior_score']}/100), "
            f"supera el objetivo global ({format_pct(global_summary['vs_goal'])}), "
            f"mantiene un perfil equilibrado y presenta fortalezas especialmente relevantes para el rol: {strengths_txt}. "
            f"Su tronco más sólido es **{candidate['strongest_trunk']}** y no muestra debilidades estructurales incompatibles con el rol."
        )

    reasons = []
    if not score_data["eligibility_checks"]["global_vs_goal"]:
        reasons.append("no supera el objetivo global")
    if not score_data["eligibility_checks"]["weakest_trunk"]:
        reasons.append(f"presenta debilidad en el tronco {candidate['weakest_trunk']}")
    if not score_data["eligibility_checks"]["critical_gaps"]:
        reasons.append("acumula carencias en indicadores críticos")
    if not score_data["eligibility_checks"]["transfer_floor"]:
        reasons.append("todavía no alcanza suficiente capacidad de transferencia/formación")

    reason_txt = ", ".join(reasons) if reasons else "queda por detrás de otros candidatos con mejor combinación global para el rol"
    return (
        f"**No propuesto como senior por ahora.** "
        f"Sus principales fortalezas son: {strengths_txt}. "
        f"No obstante, la app no lo prioriza como senior porque {reason_txt}. "
        f"Los principales focos de mejora son: {gaps_txt}."
    )


def ranking_dataframe(scored_candidates: list) -> pd.DataFrame:
    rows = []
    for idx, item in enumerate(scored_candidates, start=1):
        candidate = item["candidate"]
        score_data = item["score"]
        global_summary = candidate["global"]
        rows.append(
            {
                "Ranking": idx,
                "Nombre": candidate["name"],
                "Especie": candidate["species"],
                "Apto senior ahora": "Sí" if score_data["eligible_now"] else "No",
                "Senior Score": score_data["senior_score"],
                "Nivel global": candidate["global_level"],
                "Vs objetivo global": pct(global_summary["vs_goal"]),
                "Vs máximo global": pct(global_summary["vs_max"]),
                "Vs media BBDD": pct(global_summary["vs_bbdd"]),
                "Tronco más fuerte": candidate["strongest_trunk"],
                "Tronco más débil": candidate["weakest_trunk"],
            }
        )
    return pd.DataFrame(rows)


def detailed_export_dataframe(scored_candidates: list) -> pd.DataFrame:
    rows = []
    for item in scored_candidates:
        candidate = item["candidate"]
        score_data = item["score"]
        for _, row in candidate["indicators"].iterrows():
            rows.append(
                {
                    "Nombre": candidate["name"],
                    "Especie": candidate["species"],
                    "Tronco": row["tronco"],
                    "Indicador": row["indicator"],
                    "Score bruto": row["score_raw"],
                    "Score ponderado": row["score_weighted"],
                    "Vs objetivo": pct(row["vs_goal"]),
                    "Vs máximo": pct(row["vs_max"]),
                    "Vs media BBDD": pct(row["vs_bbdd"]),
                    "Senior Score candidato": score_data["senior_score"],
                    "Apto senior ahora": "Sí" if score_data["eligible_now"] else "No",
                }
            )
    return pd.DataFrame(rows)




def level_label(level_value):
    return EXIGENCE_LEVELS.get(level_value, {}).get("label", f"Nivel {level_value}")


def evaluate_exigence_level(score_data: dict, global_summary: dict, weakest_trunk_value: float, level: int):
    cfg = EXIGENCE_LEVELS[level]
    checks = {
        "global_vs_goal": (global_summary["vs_goal"] is not None and global_summary["vs_goal"] >= cfg["global_vs_goal"]),
        "weakest_trunk": (weakest_trunk_value >= cfg["min_trunk"]),
        "critical_gaps": (score_data["weak_critical_count"] <= cfg["max_weak_critical"]),
        "transfer_floor": (score_data["component_transfer"] >= cfg["transfer_floor"]),
        "senior_score_floor": (score_data["senior_score"] >= cfg["senior_score_floor"]),
    }
    return {"checks": checks, "eligible": all(checks.values())}


def enrich_candidate_with_exigence(item: dict):
    candidate = item["candidate"]
    score_data = item["score"]
    trunks_df = candidate["trunks"].copy()
    trunk_goal_values = [v for v in trunks_df["vs_goal"].tolist() if v is not None and not pd.isna(v)]
    weakest_trunk_value = min(trunk_goal_values) if trunk_goal_values else 0.0

    results = {}
    highest_level = 0
    for level in sorted(EXIGENCE_LEVELS.keys(), reverse=True):
        level_eval = evaluate_exigence_level(score_data, candidate["global"], weakest_trunk_value, level)
        results[level] = level_eval
        if level_eval["eligible"] and highest_level == 0:
            highest_level = level

    item["score"]["exigence_results"] = results
    item["score"]["highest_exigence_level"] = highest_level
    item["score"]["eligible_now"] = highest_level >= 1
    return item


def resolve_selected_candidate(scored_candidates: list, desired_level: int):
    eligible_at_level = [
        x for x in scored_candidates
        if x["score"]["exigence_results"][desired_level]["eligible"]
    ]
    if eligible_at_level:
        return eligible_at_level[0], desired_level, False

    for lower_level in sorted([lvl for lvl in EXIGENCE_LEVELS if lvl < desired_level], reverse=True):
        eligible_lower = [
            x for x in scored_candidates
            if x["score"]["exigence_results"][lower_level]["eligible"]
        ]
        if eligible_lower:
            return eligible_lower[0], lower_level, True

    return scored_candidates[0], 0, True


def question_default_answer(question: str, selected_candidate: dict, selected_score: dict, scored_candidates: list, applied_level: int):
    global_summary = selected_candidate["global"]
    strengths = top_strengths(selected_candidate["indicators"], top_n=3)
    gaps = main_gaps(selected_candidate["indicators"], top_n=3)
    strengths_txt = ", ".join([x["indicator"] for x in strengths]) if strengths else "no disponibles"
    gaps_txt = ", ".join([x["indicator"] for x in gaps]) if gaps else "no disponibles"
    rank_position = next((idx + 1 for idx, x in enumerate(scored_candidates) if x["candidate"]["name"] == selected_candidate["name"]), 1)
    total_candidates = len(scored_candidates)

    templates = {
        "¿El evaluado podría ser formado en áreas muy específicas o exigentes?":
            f"Sí, especialmente si se priorizan áreas donde ya muestra tracción técnica. En este caso destacan {strengths_txt}. Conviene evitar una sobreespecialización prematura en {gaps_txt} hasta consolidar base suficiente.",
        "Además de las áreas propuestas, ¿qué otras tienen más posibilidades de éxito?":
            f"Además de las áreas prioritarias, las opciones con más probabilidad de éxito son las cercanas a sus fortalezas actuales: {strengths_txt}. Son áreas donde la curva de mejora debería ser más rápida y visible.",
        "¿Está preparado para asumir mayor responsabilidad técnica a corto plazo?":
            f"Sí, con matices. El candidato está en la posición {rank_position} de {total_candidates}, con un Senior Score de {selected_score['senior_score']}/100 y un desempeño global de {format_pct(global_summary['vs_goal'])} frente a objetivo. Puede asumir mayor responsabilidad si se acompaña el desarrollo de {gaps_txt}.",
        "¿Qué debilidades limitan más su evolución en los próximos 12 meses?":
            f"Las limitaciones más claras son {gaps_txt}. Son las áreas que más condicionan una evolución rápida hacia un perfil técnico más completo y con mayor capacidad de referencia.",
        "¿Qué fortalezas son más diferenciales respecto al resto del equipo?":
            f"Las fortalezas más diferenciales son {strengths_txt}. En ellas el candidato presenta un posicionamiento comparativamente mejor frente al objetivo y/o frente a la media del equipo.",
        "¿En qué tronco debería centrarse primero el plan de mejora?":
            f"El tronco que debería priorizarse es {selected_candidate['weakest_trunk']}, porque es el que más limita la consistencia global del perfil y el que introduce mayor riesgo de desequilibrio técnico.",
        "¿Qué indicadores críticos requieren una intervención formativa inmediata?":
            f"Los indicadores críticos que piden intervención más rápida son {gaps_txt}. Son relevantes porque condicionan la robustez del perfil senior y la seguridad técnica del candidato.",
        "¿Tiene potencial real para convertirse en consultor senior en el futuro?":
            f"Sí, hay potencial real. El mejor indicador es la combinación de su Senior Score ({selected_score['senior_score']}/100), su comparación frente a objetivo ({format_pct(global_summary['vs_goal'])}) y sus fortalezas en {strengths_txt}. El ritmo dependerá de cerrar gaps en {gaps_txt}.",
        "¿Qué tipo de formación le encaja mejor: base, especializada o avanzada?":
            f"Lo más adecuado es un enfoque mixto: formación base o de consolidación en {gaps_txt}, y formación especializada en {strengths_txt} para ganar tracción y diferenciación.",
        "¿Sería recomendable exponerlo a proyectos transversales o de mayor complejidad?":
            f"Sí, pero con un criterio progresivo. Tiene sentido exponerlo a proyectos transversales relacionados con {strengths_txt}, siempre que se acompañe el refuerzo de {gaps_txt}.",
        "¿Puede actuar ya como referente parcial en alguna materia concreta?":
            f"Sí, podría actuar como referente parcial en materias cercanas a {strengths_txt}, aunque todavía no sería prudente convertirlo en referencia global si siguen abiertos los gaps en {gaps_txt}.",
        "¿Qué gap le separa con más claridad del perfil senior deseado?":
            f"El gap más claro es la combinación de {gaps_txt} con su nivel de exigencia alcanzado. Hoy su mejor ajuste encaja en {level_label(applied_level)}." if applied_level else f"El gap más claro está en {gaps_txt}, que hoy impide cumplir un estándar senior exigente.",
        "¿Qué tipo de acompañamiento interno necesitaría para acelerar su desarrollo?":
            f"El acompañamiento más útil sería mentoring técnico focalizado, revisión de casos reales y seguimiento corto sobre {gaps_txt}, aprovechando además sus fortalezas en {strengths_txt}.",
        "¿Qué fortalezas conviene consolidar para convertirlas en especialización?":
            f"Conviene consolidar especialmente {strengths_txt}, porque son las áreas con más probabilidad de convertirse en una verdadera especialización técnica.",
        "¿Hay riesgo de sobreestimar su perfil por desequilibrios entre troncos?":
            f"Sí, el principal riesgo de sobreestimación aparece cuando un buen rendimiento global convive con debilidad relativa en {selected_candidate['weakest_trunk']}. Por eso la lectura debe hacerse siempre por troncos, no solo con la nota total.",
        "¿Qué impacto tendría priorizar su formación en herramientas y análisis de datos?":
            f"Priorizar herramientas y análisis de datos tendría un impacto alto si esa área está entre sus gaps, porque mejora capacidad de diagnóstico, estructuración de criterio y transferencia al equipo.",
        "¿Qué combinación de formación interna y externa parece más eficiente para este técnico?":
            f"La combinación más eficiente sería formación interna muy dirigida en {gaps_txt} y formación externa selectiva en {strengths_txt}, reservando lo externo para especialización o credenciales con mayor valor diferencial.",
        "¿Está más cerca de un perfil especialista o de un perfil generalista sólido?":
            f"Hoy parece más cercano a {'un perfil especialista' if strengths and gaps else 'un perfil generalista sólido'}, con mayor tracción en {strengths_txt} y necesidad de equilibrar {gaps_txt}.",
        "¿Qué argumentos objetivos justifican que no sea el elegido ahora mismo?":
            f"Los argumentos objetivos son su posición relativa en el ranking ({rank_position}/{total_candidates}), el nivel de exigencia que realmente alcanza ({level_label(selected_score.get('highest_exigence_level', 0)) if selected_score.get('highest_exigence_level', 0) else 'ninguno'}) y los gaps en {gaps_txt}.",
        "¿Cuál sería el siguiente paso más razonable en su itinerario profesional?":
            f"El siguiente paso más razonable es un plan de 6–12 meses centrado en {gaps_txt}, mientras se refuerzan y visibilizan sus fortalezas en {strengths_txt} para consolidar un perfil más completo.",
    }
    return templates.get(
        question,
        f"Lectura preliminar: el candidato destaca en {strengths_txt}, pero necesita seguir desarrollando {gaps_txt} para ganar solidez y proyección."
    )


def ensure_question_state_for_selected(candidate_name: str):
    owner = st.session_state.get("qa_owner")
    if owner != candidate_name:
        st.session_state["qa_items"] = []
        st.session_state["qa_owner"] = candidate_name


def add_question_item(question: str, selected_candidate: dict, selected_score: dict, scored_candidates: list, applied_level: int):
    qa_items = st.session_state.get("qa_items", [])
    if any(item["question"] == question for item in qa_items):
        return
    qa_items.append(
        {
            "question": question,
            "include": True,
            "answer": question_default_answer(question, selected_candidate, selected_score, scored_candidates, applied_level),
        }
    )
    st.session_state["qa_items"] = qa_items


def question_dataframe_from_state():
    qa_items = st.session_state.get("qa_items", [])
    rows = []
    for item in qa_items:
        rows.append(
            {
                "Pregunta": item["question"],
                "Incluir en informe": "Sí" if item["include"] else "No",
                "Respuesta": item["answer"],
            }
        )
    return pd.DataFrame(rows)


def build_global_summary_report(selected_candidate: dict, selected_score: dict, scored_candidates: list, desired_level: int, applied_level: int, level_fallback: bool):
    global_summary = selected_candidate["global"]
    selected_name = selected_candidate["name"]
    strengths = top_strengths(selected_candidate["indicators"], top_n=3)
    gaps = main_gaps(selected_candidate["indicators"], top_n=3)
    strengths_txt = ", ".join([x["indicator"] for x in strengths]) if strengths else "sin fortalezas claramente diferenciales"
    gaps_txt = ", ".join([x["indicator"] for x in gaps]) if gaps else "sin gaps claramente identificados"
    apt_now = sum(1 for x in scored_candidates if x["score"]["exigence_results"].get(applied_level, {}).get("eligible", False)) if applied_level else 0

    lines = []
    lines.append(f"Se han analizado {len(scored_candidates)} candidatos para identificar el perfil más adecuado como consultor senior.")
    lines.append(f"El candidato propuesto es {selected_name}, con un Senior Score de {selected_score['senior_score']}/100.")
    lines.append(f"Su comparación global frente al objetivo es {format_pct(global_summary['vs_goal'])} y frente a la media BBDD es {format_pct(global_summary['vs_bbdd'])}.")
    lines.append(f"El tronco más fuerte es {selected_candidate['strongest_trunk']} y el tronco más débil es {selected_candidate['weakest_trunk']}.")
    lines.append(f"Sus fortalezas más relevantes son {strengths_txt}.")
    lines.append(f"Los principales focos de mejora son {gaps_txt}.")
    if applied_level:
        lines.append(f"La decisión final se ha tomado con {level_label(applied_level)}.")
        if level_fallback:
            lines.append(f"No hubo candidato que cumpliera el nivel de exigencia {desired_level}, por lo que la app aplicó el siguiente nivel disponible.")
        else:
            lines.append(f"Sí existe al menos un candidato que cumple el nivel de exigencia seleccionado inicialmente.")
        lines.append(f"En el nivel finalmente aplicado hay {apt_now} candidato(s) apto(s).")
    else:
        lines.append("Ningún candidato cumple todavía un estándar senior formal en los niveles 3, 2 o 1.")
        lines.append("Aun así, la app propone el mejor perfil relativo disponible para orientar la decisión y el plan de desarrollo.")
    lines.append("La elección no depende solo de la nota global, sino del equilibrio entre troncos, los indicadores críticos, la capacidad de transferencia y la ventaja respecto al equipo.")
    lines.append("La recomendación debe leerse como una base objetiva para decidir, comunicar y construir el siguiente plan de desarrollo técnico.")
    return "\n".join(lines)


def build_report_with_questions(base_report: str):
    qa_items = [item for item in st.session_state.get("qa_items", []) if item.get("include")]
    if not qa_items:
        return base_report

    blocks = [base_report, "", "Preguntas complementarias incorporadas al informe:"]
    for idx, item in enumerate(qa_items, start=1):
        blocks.append(f"{idx}. {item['question']}")
        blocks.append(item["answer"])
        blocks.append("")
    return "\n".join(blocks).strip()


def build_excel_report(scored_candidates: list, report_text: str = "", qa_df: pd.DataFrame | None = None):
    ranking_df = ranking_dataframe(scored_candidates)
    detail_df = detailed_export_dataframe(scored_candidates)
    method_df = CRITERIA_TABLE.copy()

    bio = io.BytesIO()
    with pd.ExcelWriter(bio, engine="xlsxwriter") as writer:
        ranking_df.to_excel(writer, sheet_name="Ranking", index=False)
        detail_df.to_excel(writer, sheet_name="Detalle", index=False)
        method_df.to_excel(writer, sheet_name="Criterios", index=False)

        if qa_df is not None and not qa_df.empty:
            qa_df.to_excel(writer, sheet_name="Preguntas", index=False)

        if report_text:
            report_df = pd.DataFrame({"Informe final": report_text.split("\n")})
            report_df.to_excel(writer, sheet_name="Informe", index=False)

    bio.seek(0)
    return bio.getvalue()



# =========================================================
# APP
# =========================================================

initialize_app_state()

st.title("Senior Technical Consultant Selector")
st.caption(
    "Sube entre 2 y 10 evaluaciones individuales y la app generará un ranking completo de candidatos "
    "para consultor senior, con argumentos técnicos, criterios de selección transparentes y un módulo de preguntas editables."
)

with st.expander("Criterios de selección del senior", expanded=True):
    st.markdown(CRITERIA_TEXT)
    st.table(CRITERIA_TABLE)
    st.markdown("**Niveles de exigencia disponibles**")
    st.table(EXIGENCE_TABLE)

st.info(
    "Importante: la app funciona mejor si los ficheros se han abierto y guardado previamente en Excel "
    "con las fórmulas actualizadas."
)

control_col1, control_col2, control_col3 = st.columns([2.5, 1.3, 1.2])
with control_col1:
    desired_level = st.select_slider(
        "Nivel de exigencia para la elección del senior",
        options=[1, 2, 3],
        value=3,
        format_func=lambda x: level_label(x),
    )
with control_col2:
    st.write("")
    st.write("")
    if st.button("Nuevo análisis", use_container_width=True):
        reset_app_state()
        st.rerun()
with control_col3:
    st.write("")
    st.write("")
    if st.button("Limpiar preguntas", use_container_width=True):
        st.session_state["qa_items"] = []
        st.rerun()

uploaded_files = st.file_uploader(
    "Sube de 2 a 10 evaluaciones (.xlsm o .xlsx)",
    type=["xlsm", "xlsx"],
    accept_multiple_files=True,
    help="Cada archivo debe corresponder a una evaluación individual completa.",
    key=f"uploader_{st.session_state['uploader_key']}",
)

if not uploaded_files:
    st.stop()

if len(uploaded_files) < 2:
    st.warning("Necesitas subir al menos 2 evaluaciones para construir un ranking comparativo.")
    st.stop()

if len(uploaded_files) > 10:
    st.warning("Por ahora la app está pensada para un máximo de 10 evaluaciones por análisis.")
    st.stop()

parsed_candidates = []
errors = []

for file in uploaded_files:
    try:
        candidate = parse_candidate(file)
        score_data = score_candidate(candidate)
        parsed_candidates.append({"candidate": candidate, "score": score_data})
    except Exception as exc:
        errors.append(f"{file.name}: {exc}")

if errors:
    st.error("Algunos archivos no se han podido procesar:")
    for err in errors:
        st.write(f"- {err}")

if not parsed_candidates:
    st.stop()

parsed_candidates = [enrich_candidate_with_exigence(item) for item in parsed_candidates]
parsed_candidates = sorted(
    parsed_candidates,
    key=lambda x: (
        x["score"].get("highest_exigence_level", 0),
        x["score"]["senior_score"],
        x["candidate"]["global"]["vs_goal"] or 0,
    ),
    reverse=True,
)

selected_item, applied_level, level_fallback = resolve_selected_candidate(parsed_candidates, desired_level)
selected_candidate = selected_item["candidate"]
selected_score = selected_item["score"]

ensure_question_state_for_selected(selected_candidate["name"])

st.subheader("Conclusión ejecutiva")

if applied_level > 0 and selected_score["exigence_results"][applied_level]["eligible"]:
    if level_fallback and applied_level != desired_level:
        st.warning(
            f"No hay candidato que cumpla {level_label(desired_level)}. "
            f"La app propone a **{selected_candidate['name']}** usando {level_label(applied_level)} "
            f"(Senior Score {selected_score['senior_score']}/100)."
        )
    else:
        st.success(
            f"**Candidato propuesto para consultor senior: {selected_candidate['name']}** "
            f"(Senior Score {selected_score['senior_score']}/100, {level_label(applied_level)})."
        )
else:
    st.warning(
        f"**No aparece un senior plenamente consolidado en este grupo.** "
        f"El mejor perfil actual es **{selected_candidate['name']}** "
        f"(Senior Score {selected_score['senior_score']}/100), pero todavía no cumple todo el estándar senior."
    )

st.markdown(candidate_argument(selected_candidate, selected_score, selected_name=selected_candidate["name"]))

st.subheader("Ranking final de candidatos")
ranking_df = ranking_dataframe(parsed_candidates)
ranking_df["Nivel máximo alcanzado"] = ranking_df["Nombre"].map(
    {
        item["candidate"]["name"]: level_label(item["score"]["highest_exigence_level"]) if item["score"]["highest_exigence_level"] else "Ninguno"
        for item in parsed_candidates
    }
)
st.dataframe(ranking_df, use_container_width=True, hide_index=True)

col1, col2, col3, col4, col5 = st.columns(5)
with col1:
    st.metric("Nº candidatos", len(parsed_candidates))
with col2:
    st.metric("Aptos al nivel aplicado", sum(1 for x in parsed_candidates if applied_level > 0 and x["score"]["exigence_results"][applied_level]["eligible"]))
with col3:
    st.metric("Mejor Senior Score", f"{parsed_candidates[0]['score']['senior_score']}/100")
with col4:
    st.metric("Nivel aplicado", level_label(applied_level) if applied_level else "Sin nivel")
with col5:
    st.metric("Candidato propuesto", selected_candidate["name"])

st.markdown("**Comparación visual del ranking final**")
st.plotly_chart(build_bar_chart(ranking_df), use_container_width=True)

st.markdown("**Mapa de fortalezas y debilidades de los candidatos**")
st.caption(
    "El gráfico de araña compara los 5 componentes del Senior Score. Cuanto más lejos del centro, mejor posicionamiento en ese componente."
)
st.plotly_chart(build_radar_chart(parsed_candidates), use_container_width=True)

st.subheader("Preguntas frecuentes para enriquecer el informe")
st.caption(
    "Selecciona una de las 20 preguntas propuestas, genera una respuesta editable y decide si quieres incorporarla al informe final."
)

qcol1, qcol2 = st.columns([3.5, 1.2])
with qcol1:
    selected_question = st.selectbox(
        "Pregunta a añadir",
        POSSIBLE_QUESTIONS,
        key="question_selector",
    )
with qcol2:
    st.write("")
    st.write("")
    if st.button("Añadir pregunta", use_container_width=True):
        add_question_item(selected_question, selected_candidate, selected_score, parsed_candidates, applied_level)
        st.rerun()

qa_items = st.session_state.get("qa_items", [])
if qa_items:
    remove_index = None
    for idx, item in enumerate(qa_items):
        with st.expander(f"Pregunta {idx + 1}: {item['question']}", expanded=(idx == 0)):
            c1, c2 = st.columns([1.2, 1])
            include_key = f"qa_include_{idx}"
            answer_key = f"qa_answer_{idx}"
            st.session_state.setdefault(include_key, item["include"])
            st.session_state.setdefault(answer_key, item["answer"])

            include_value = c1.checkbox("Incorporar al informe", value=st.session_state[include_key], key=include_key)
            if c2.button("Eliminar pregunta", key=f"remove_qa_{idx}", use_container_width=True):
                remove_index = idx

            answer_value = st.text_area(
                "Respuesta editable",
                value=st.session_state[answer_key],
                key=answer_key,
                height=150,
            )

            st.session_state["qa_items"][idx]["include"] = include_value
            st.session_state["qa_items"][idx]["answer"] = answer_value

    if remove_index is not None:
        st.session_state["qa_items"].pop(remove_index)
        st.rerun()
else:
    st.info("Todavía no has añadido preguntas complementarias al informe.")

st.subheader("Detalle y argumentos por candidato")

for item in parsed_candidates:
    candidate = item["candidate"]
    score_data = item["score"]
    global_summary = candidate["global"]

    title = (
        f"{candidate['name']} · Senior Score {score_data['senior_score']}/100 · "
        f"{level_label(score_data['highest_exigence_level']) if score_data['highest_exigence_level'] else 'Sin nivel alcanzado'}"
    )
    with st.expander(title, expanded=(candidate["name"] == selected_candidate["name"])):
        left, right = st.columns([1.05, 1])

        with left:
            st.markdown(candidate_argument(candidate, score_data, selected_name=selected_candidate["name"]))

            components_df = pd.DataFrame(
                [
                    {"Componente": "Rendimiento global", "Score": score_data["component_global"]},
                    {"Componente": "Equilibrio entre troncos", "Score": score_data["component_balance"]},
                    {"Componente": "Indicadores críticos", "Score": score_data["component_critical"]},
                    {"Componente": "Transferencia y formación", "Score": score_data["component_transfer"]},
                    {"Componente": "Ventaja frente a la media", "Score": score_data["component_team_advantage"]},
                ]
            )
            st.markdown("**Desglose del ranking final**")
            st.dataframe(components_df, use_container_width=True, hide_index=True)

            level_rows = []
            for level in sorted(EXIGENCE_LEVELS.keys(), reverse=True):
                checks = score_data["exigence_results"][level]["checks"]
                level_rows.append(
                    {
                        "Nivel": level_label(level),
                        "Apto": "Sí" if score_data["exigence_results"][level]["eligible"] else "No",
                        "Objetivo global": "Sí" if checks["global_vs_goal"] else "No",
                        "Peor tronco": "Sí" if checks["weakest_trunk"] else "No",
                        "Críticos": "Sí" if checks["critical_gaps"] else "No",
                        "Transferencia": "Sí" if checks["transfer_floor"] else "No",
                        "Senior Score": "Sí" if checks["senior_score_floor"] else "No",
                    }
                )
            st.markdown("**Cumplimiento por nivel de exigencia**")
            st.dataframe(pd.DataFrame(level_rows), use_container_width=True, hide_index=True)

        with right:
            metric_cols = st.columns(2)
            metric_cols[0].metric("Vs objetivo global", format_pct(global_summary["vs_goal"]))
            metric_cols[1].metric("Vs máximo global", format_pct(global_summary["vs_max"]))
            metric_cols[0].metric("Vs media BBDD", format_pct(global_summary["vs_bbdd"]))
            metric_cols[1].metric("Nivel global", candidate["global_level"])
            metric_cols[0].metric("Tronco más fuerte", candidate["strongest_trunk"])
            metric_cols[1].metric("Tronco más débil", candidate["weakest_trunk"])

            trunks_display = candidate["trunks"][["tronco", "vs_goal", "vs_max", "vs_bbdd"]].copy()
            trunks_display["vs_goal"] = trunks_display["vs_goal"].apply(format_pct)
            trunks_display["vs_max"] = trunks_display["vs_max"].apply(format_pct)
            trunks_display["vs_bbdd"] = trunks_display["vs_bbdd"].apply(format_pct)
            st.markdown("**Comparación por troncos**")
            st.dataframe(trunks_display, use_container_width=True, hide_index=True)

        indicator_display = candidate["indicators"][["tronco", "indicator", "score_raw", "weight", "vs_goal", "vs_max", "vs_bbdd"]].copy()
        indicator_display["vs_goal"] = indicator_display["vs_goal"].apply(format_pct)
        indicator_display["vs_max"] = indicator_display["vs_max"].apply(format_pct)
        indicator_display["vs_bbdd"] = indicator_display["vs_bbdd"].apply(format_pct)
        indicator_display = indicator_display.rename(
            columns={
                "tronco": "Tronco",
                "indicator": "Indicador",
                "score_raw": "Score",
                "weight": "Peso",
                "vs_goal": "Vs objetivo",
                "vs_max": "Vs máximo",
                "vs_bbdd": "Vs media BBDD",
            }
        )
        st.markdown("**Detalle por indicador**")
        st.dataframe(indicator_display, use_container_width=True, hide_index=True)

st.subheader("Informe final editable")
base_report = build_global_summary_report(selected_candidate, selected_score, parsed_candidates, desired_level, applied_level, level_fallback)
full_report = build_report_with_questions(base_report)

if not st.session_state.get("report_text") or st.button("Regenerar informe con las preguntas seleccionadas"):
    st.session_state["report_text"] = full_report

st.text_area(
    "Texto final del informe",
    key="report_text",
    height=380,
    help="Puedes editar libremente este texto antes de exportarlo.",
)

st.subheader("Descargas")
qa_df = question_dataframe_from_state()
export_bytes = build_excel_report(parsed_candidates, report_text=st.session_state["report_text"], qa_df=qa_df)

d1, d2, d3 = st.columns(3)
with d1:
    st.download_button(
        "Descargar ranking y detalle en Excel",
        data=export_bytes,
        file_name="senior_consultant_ranking.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )
with d2:
    json_payload = {
        item["candidate"]["name"]: {
            "species": item["candidate"]["species"],
            "senior_score": item["score"]["senior_score"],
            "highest_exigence_level": item["score"]["highest_exigence_level"],
            "global": item["candidate"]["global"],
        }
        for item in parsed_candidates
    }
    st.download_button(
        "Descargar resumen JSON",
        data=json.dumps(json_payload, ensure_ascii=False, indent=2, default=str).encode("utf-8"),
        file_name="senior_consultant_ranking.json",
        mime="application/json",
        use_container_width=True,
    )
with d3:
    st.download_button(
        "Descargar informe en TXT",
        data=st.session_state["report_text"].encode("utf-8"),
        file_name="informe_final_senior.txt",
        mime="text/plain",
        use_container_width=True,
    )
